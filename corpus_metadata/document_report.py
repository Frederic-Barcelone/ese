"""
Biomedical Entity Extractor - Database to Excel Exporter
Reads extraction results from SQLite database and exports to Excel
"""

import sqlite3
import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class BiomedicaDBToExcel:
    """Exports biomedical entity extraction results from database to Excel"""
    
    def __init__(self, db_path='./corpus_db/extraction_results.db'):
        self.db_path = db_path
        self.conn = None
        
    def connect(self):
        """Connect to the SQLite database"""
        if not Path(self.db_path).exists():
            raise FileNotFoundError(f"Database not found: {self.db_path}")
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        
    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()
            
    def get_all_documents(self):
        """Retrieve all documents from database"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM documents ORDER BY created_at DESC")
        return cursor.fetchall()
    
    def get_entities_for_document(self, doc_id):
        """Retrieve all entities for a specific document"""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT * FROM entities 
            WHERE doc_id = ? 
            ORDER BY entity_type, normalized
        """, (doc_id,))
        return cursor.fetchall()
    
    def get_abbreviations_for_document(self, doc_id):
        """Retrieve abbreviation map for a specific document"""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT * FROM abbreviations 
            WHERE doc_id = ? 
            ORDER BY abbreviation
        """, (doc_id,))
        return cursor.fetchall()
    
    def export_to_excel(self, output_path='biomedical_extraction_results.xlsx'):
        """Export all data to Excel with multiple sheets"""
        self.connect()
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create sheets
            self._create_summary_sheet(wb)
            self._create_entities_sheet(wb)
            self._create_diseases_sheet(wb)
            self._create_drugs_sheet(wb)
            self._create_abbreviations_sheet(wb)
            self._create_documents_sheet(wb)
            
            # Save workbook
            wb.save(output_path)
            print(f"✓ Excel file created: {output_path}")
            
            return output_path
            
        finally:
            self.close()
    
    def _create_summary_sheet(self, wb):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Header styling
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws['A1'] = "Biomedical Entity Extraction - Summary Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:D2')
        
        # Get statistics
        cursor = self.conn.cursor()
        
        # Document count
        cursor.execute("SELECT COUNT(*) FROM documents")
        doc_count = cursor.fetchone()[0]
        
        # Entity counts by type
        cursor.execute("""
            SELECT entity_type, COUNT(*) as count 
            FROM entities 
            GROUP BY entity_type
        """)
        entity_counts = dict(cursor.fetchall())
        
        # Total entities
        cursor.execute("SELECT COUNT(*) FROM entities")
        total_entities = cursor.fetchone()[0]
        
        # Abbreviations count
        cursor.execute("SELECT COUNT(DISTINCT abbreviation) FROM abbreviations")
        abbrev_count = cursor.fetchone()[0]
        
        # Average confidence by type
        cursor.execute("""
            SELECT entity_type, ROUND(AVG(confidence), 3) as avg_conf
            FROM entities 
            GROUP BY entity_type
        """)
        avg_confidences = dict(cursor.fetchall())
        
        # Write statistics
        row = 4
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        
        stats = [
            ("Total Documents", doc_count),
            ("Total Entities Extracted", total_entities),
            ("", ""),
            ("DISEASE Entities", entity_counts.get('DISEASE', 0)),
            ("DRUG Entities", entity_counts.get('DRUG', 0)),
            ("ABBREVIATION Entities", entity_counts.get('ABBREVIATION', 0)),
            ("Unique Abbreviations", abbrev_count),
            ("", ""),
            ("Avg Confidence - DISEASE", avg_confidences.get('DISEASE', 0)),
            ("Avg Confidence - DRUG", avg_confidences.get('DRUG', 0)),
            ("Avg Confidence - ABBREVIATION", avg_confidences.get('ABBREVIATION', 0)),
        ]
        
        for metric, value in stats:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            if metric:
                ws[f'A{row}'].font = Font(bold=True)
        
        # Format columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        
    def _create_entities_sheet(self, wb):
        """Create sheet with all entities"""
        ws = wb.create_sheet("All Entities")
        
        # Headers
        headers = [
            "Doc ID", "Entity Type", "Text", "Normalized", 
            "Confidence", "Status", "Negated", "Hypothetical", 
            "Conditional", "Investigational", "Identifiers", "Evidence"
        ]
        
        self._write_headers(ws, headers)
        
        # Get all entities
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT e.*, d.filename 
            FROM entities e
            LEFT JOIN documents d ON e.doc_id = d.doc_id
            ORDER BY e.doc_id, e.entity_type, e.confidence DESC
        """)
        
        row = 2
        for entity in cursor.fetchall():
            ws[f'A{row}'] = entity['doc_id']
            ws[f'B{row}'] = entity['entity_type']
            ws[f'C{row}'] = entity['text']
            ws[f'D{row}'] = entity['normalized']
            ws[f'E{row}'] = round(entity['confidence'], 3)
            ws[f'F{row}'] = entity['status']
            
            # Parse context flags
            context_flags = json.loads(entity['context_flags']) if entity['context_flags'] else {}
            ws[f'G{row}'] = "Yes" if context_flags.get('negated') else "No"
            ws[f'H{row}'] = "Yes" if context_flags.get('hypothetical') else "No"
            ws[f'I{row}'] = "Yes" if context_flags.get('conditional') else "No"
            ws[f'J{row}'] = "Yes" if entity['investigational'] else "No"
            
            # Parse identifiers
            identifiers = json.loads(entity['identifiers']) if entity['identifiers'] else {}
            ws[f'K{row}'] = json.dumps(identifiers) if identifiers else ""
            
            ws[f'L{row}'] = entity['evidence'][:100] if entity['evidence'] else ""
            
            row += 1
        
        self._format_sheet(ws, len(headers))
        
    def _create_diseases_sheet(self, wb):
        """Create sheet focused on disease entities"""
        ws = wb.create_sheet("Diseases")
        
        headers = [
            "Disease Name", "Normalized", "Count", "Avg Confidence",
            "Documents", "Identifiers", "Sample Evidence"
        ]
        
        self._write_headers(ws, headers)
        
        # Get disease statistics
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT 
                normalized,
                text,
                COUNT(*) as mention_count,
                ROUND(AVG(confidence), 3) as avg_conf,
                GROUP_CONCAT(DISTINCT doc_id) as docs,
                identifiers,
                evidence
            FROM entities
            WHERE entity_type = 'DISEASE'
            GROUP BY normalized
            ORDER BY mention_count DESC, avg_conf DESC
        """)
        
        row = 2
        for disease in cursor.fetchall():
            ws[f'A{row}'] = disease['text']
            ws[f'B{row}'] = disease['normalized']
            ws[f'C{row}'] = disease['mention_count']
            ws[f'D{row}'] = disease['avg_conf']
            
            # Document list
            docs = disease['docs'].split(',') if disease['docs'] else []
            ws[f'E{row}'] = len(docs)
            
            # Identifiers
            identifiers = json.loads(disease['identifiers']) if disease['identifiers'] else {}
            ws[f'F{row}'] = json.dumps(identifiers) if identifiers else ""
            
            ws[f'G{row}'] = disease['evidence'][:100] if disease['evidence'] else ""
            
            row += 1
        
        self._format_sheet(ws, len(headers))
        
    def _create_drugs_sheet(self, wb):
        """Create sheet focused on drug entities"""
        ws = wb.create_sheet("Drugs")
        
        headers = [
            "Drug Name", "Normalized (Generic)", "Count", "Avg Confidence",
            "Investigational", "Documents", "Identifiers", "Sample Evidence"
        ]
        
        self._write_headers(ws, headers)
        
        # Get drug statistics
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT 
                text,
                normalized,
                COUNT(*) as mention_count,
                ROUND(AVG(confidence), 3) as avg_conf,
                MAX(investigational) as is_investigational,
                GROUP_CONCAT(DISTINCT doc_id) as docs,
                identifiers,
                evidence
            FROM entities
            WHERE entity_type = 'DRUG'
            GROUP BY normalized
            ORDER BY mention_count DESC, avg_conf DESC
        """)
        
        row = 2
        for drug in cursor.fetchall():
            ws[f'A{row}'] = drug['text']
            ws[f'B{row}'] = drug['normalized']
            ws[f'C{row}'] = drug['mention_count']
            ws[f'D{row}'] = drug['avg_conf']
            ws[f'E{row}'] = "Yes" if drug['is_investigational'] else "No"
            
            # Document list
            docs = drug['docs'].split(',') if drug['docs'] else []
            ws[f'F{row}'] = len(docs)
            
            # Identifiers
            identifiers = json.loads(drug['identifiers']) if drug['identifiers'] else {}
            ws[f'G{row}'] = json.dumps(identifiers) if identifiers else ""
            
            ws[f'H{row}'] = drug['evidence'][:100] if drug['evidence'] else ""
            
            row += 1
        
        self._format_sheet(ws, len(headers))
        
    def _create_abbreviations_sheet(self, wb):
        """Create sheet with abbreviation mappings"""
        ws = wb.create_sheet("Abbreviations")
        
        headers = ["Abbreviation", "Expansion", "Confidence", "Documents", "Context"]
        
        self._write_headers(ws, headers)
        
        # Get abbreviations
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT 
                abbreviation,
                expansion,
                ROUND(AVG(confidence), 3) as avg_conf,
                COUNT(DISTINCT doc_id) as doc_count,
                GROUP_CONCAT(DISTINCT doc_id) as docs
            FROM abbreviations
            GROUP BY abbreviation, expansion
            ORDER BY abbreviation
        """)
        
        row = 2
        for abbr in cursor.fetchall():
            ws[f'A{row}'] = abbr['abbreviation']
            ws[f'B{row}'] = abbr['expansion']
            ws[f'C{row}'] = abbr['avg_conf']
            ws[f'D{row}'] = abbr['doc_count']
            ws[f'E{row}'] = abbr['docs']
            
            row += 1
        
        self._format_sheet(ws, len(headers))
        
    def _create_documents_sheet(self, wb):
        """Create sheet with document information"""
        ws = wb.create_sheet("Documents")
        
        headers = [
            "Doc ID", "Filename", "Entity Count", 
            "Disease Count", "Drug Count", "Abbrev Count",
            "Warnings", "Created At"
        ]
        
        self._write_headers(ws, headers)
        
        documents = self.get_all_documents()
        
        row = 2
        for doc in documents:
            # Get entity counts
            cursor = self.conn.cursor()
            cursor.execute("""
                SELECT COUNT(*) FROM entities WHERE doc_id = ?
            """, (doc['doc_id'],))
            entity_count = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COUNT(*) FROM entities 
                WHERE doc_id = ? AND entity_type = 'DISEASE'
            """, (doc['doc_id'],))
            disease_count = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COUNT(*) FROM entities 
                WHERE doc_id = ? AND entity_type = 'DRUG'
            """, (doc['doc_id'],))
            drug_count = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COUNT(DISTINCT abbreviation) FROM abbreviations 
                WHERE doc_id = ?
            """, (doc['doc_id'],))
            abbrev_count = cursor.fetchone()[0]
            
            ws[f'A{row}'] = doc['doc_id']
            ws[f'B{row}'] = doc['filename']
            ws[f'C{row}'] = entity_count
            ws[f'D{row}'] = disease_count
            ws[f'E{row}'] = drug_count
            ws[f'F{row}'] = abbrev_count
            
            # Parse warnings
            warnings = json.loads(doc['warnings']) if doc['warnings'] else []
            ws[f'G{row}'] = "; ".join(warnings) if warnings else ""
            
            ws[f'H{row}'] = doc['created_at']
            
            row += 1
        
        self._format_sheet(ws, len(headers))
        
    def _write_headers(self, ws, headers):
        """Write and style header row"""
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    def _format_sheet(self, ws, num_columns):
        """Apply formatting to sheet"""
        # Auto-adjust column widths
        for col_num in range(1, num_columns + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                min_col=1, max_col=num_columns):
            for cell in row:
                cell.border = thin_border


def main():
    """Main execution function"""
    db_path = './corpus_db/extraction_results.db'
    output_path = f'biomedical_extraction_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    print("Biomedical Entity Extraction - Database to Excel Exporter")
    print("=" * 60)
    print(f"Database: {db_path}")
    print(f"Output: {output_path}")
    print()
    
    try:
        exporter = BiomedicaDBToExcel(db_path)
        result_path = exporter.export_to_excel(output_path)
        
        print()
        print("=" * 60)
        print(f"✓ Export completed successfully!")
        print(f"✓ File saved: {result_path}")
        print()
        print("Sheets created:")
        print("  • Summary - Overall statistics")
        print("  • All Entities - Complete entity list")
        print("  • Diseases - Disease-specific analysis")
        print("  • Drugs - Drug-specific analysis")
        print("  • Abbreviations - Abbreviation mappings")
        print("  • Documents - Document metadata")
        
    except FileNotFoundError as e:
        print(f"✗ Error: {e}")
        print("Please ensure the database file exists at the specified path.")
    except Exception as e:
        print(f"✗ Error during export: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()